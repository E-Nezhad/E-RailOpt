import numpy as np
import pandas as pd
from propulsionModel import get_input_segment, totalResistanceForces, tractive_power
import string
from openpyxl.utils import get_column_letter


def compute_energy_matrix_with_stops(stations_m, P):
    n = len(stations_m)
    energy_matrix = np.zeros((n, n))
    regen_buffer = 0

    for i in range(n):
        for j in range(n):
            if j <= i:
                continue  # only compute forward direction
            start_m = stations_m[i]
            end_m = stations_m[j]

            print(f"Simulating from station {i} ({start_m} m) to station {j} ({end_m} m)...")

            # Get simulation results
            acc, vel, grade, curv, dist = get_input_segment(P, start_m, end_m)
            resist = totalResistanceForces(vel, grade, curv)
            _, _, _, _, _, total_energy_kwh, regen_buffer = tractive_power(acc, resist, vel, regen_buffer)

            energy_matrix[i][j] = total_energy_kwh

    return energy_matrix

def compute_time_matrix(stations_m, P):
    n = len(stations_m)
    time_matrix = np.zeros((n, n))

    for i in range(n):
        for j in range(n):
            if i == j:
                time_matrix[i, j] = 0.0
            elif j > i:
                # Simulate train from station i to station j
                _, _, _, _, distance_seconds = get_input_segment(P, stations_m[i], stations_m[j])
                # The simulation time is length of distance_seconds array * dt (1 sec)
                travel_time = len(distance_seconds) - 1  # seconds
                time_matrix[i, j] = travel_time
            else:
                # For j < i, travel time can be same as reverse or NaN if one-way
                time_matrix[i, j] = np.nan  # or replicate i-j if bidirectional

    return time_matrix


def export_energy_matrix_to_excel(matrix, stations_m, output_path="energy_matrix.xlsx"):
    # Create station labels like "Station A (3500 m)"
    labels = [f"Station {string.ascii_uppercase[i]} ({int(s)} m)" for i, s in enumerate(stations_m)]
    
    df = pd.DataFrame(matrix, index=labels, columns=labels)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Energy Matrix")

        sheet = writer.sheets["Energy Matrix"]
        book = writer.book  # just read, don't assign

        # Autofit column widths based on content
        for column_cells in sheet.columns:
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells
            )
            col_letter = get_column_letter(column_cells[0].column)
            sheet.column_dimensions[col_letter].width = max_length + 2  # padding

    print(f"\n✅ Energy matrix saved to '{output_path}'")

def export_matrices_to_excel(energy_matrix, time_matrix, stations_m, output_path="matrices.xlsx"):
    labels = [f"Station {string.ascii_uppercase[i]} ({int(s)} m)" for i, s in enumerate(stations_m)]
    
    # Create DataFrames
    df_energy = pd.DataFrame(energy_matrix, index=labels, columns=labels)
    df_time = pd.DataFrame(time_matrix, index=labels, columns=labels)
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Write energy matrix
        df_energy.to_excel(writer, sheet_name="Energy Matrix")
        sheet_energy = writer.sheets["Energy Matrix"]
        
        # Autofit energy matrix columns
        for column_cells in sheet_energy.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            col_letter = get_column_letter(column_cells[0].column)
            sheet_energy.column_dimensions[col_letter].width = max_length + 2
        
        # Write time matrix
        df_time.to_excel(writer, sheet_name="Time Matrix")
        sheet_time = writer.sheets["Time Matrix"]
        
        # Autofit time matrix columns
        for column_cells in sheet_time.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            col_letter = get_column_letter(column_cells[0].column)
            sheet_time.column_dimensions[col_letter].width = max_length + 2

    print(f"\n✅ Energy and time matrices saved to '{output_path}'")

def main():
    # Example station chainages (in meters)
    stations_km = [0, 1000, 2000, 3000, 3500, 4000]  # replace with your actual station chainages
    stations_m = [s * 1.0 for s in stations_km]  # ensure float type

    P = 3000 * 1000  # tractive power in Watts

    energy_matrix = compute_energy_matrix_with_stops(stations_m, P)
    time_matrix = compute_time_matrix(stations_m, P)
    export_matrices_to_excel(energy_matrix,time_matrix, stations_m)

if __name__ == "__main__":
    main()

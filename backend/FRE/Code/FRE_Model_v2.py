# -*- coding: utf-8 -*-
"""
Created on Fri Oct 18 16:06:37 2024

@author: z3547138
"""

import numpy as np
import openpyxl
import os
import pandas as pd
import pyomo.environ as pe
import pyomo.opt
import random
import scipy as sp
import shutil
import time
import timeit
import xlrd
import xlsxwriter
import zipfile
from datetime import datetime, date
from openpyxl import Workbook
from pathlib import Path
from pyomo.environ import *
from xls2xlsx import XLS2XLSX
from FRE_Functions import *

'''This file:
    1. reads random instance file from FRE/Data/Random, 
    2. solves the problem by integrated model model_change_rate_v2_PLARec,
    3. export results to FRE/Results/Random Instances'''

### sophia changed the below from range(21,31) to range(1, 2)
id_instance_random_list = list(range(1,2))
'''algorithm parameters'''
alg = 'PLARec'
M, epsilon = 1000, 0.000001
gap_op = 0.1 # tolerance gap when solving Model_change_rate_v2_PLARec 
time_limit_op = 1800 # time limit to solve Model_change_rate_v2_PLARec
mysolver = 'gurobi'
buffer_soc_estimate = 1.13  # in constraints soc_time_charge_rule_a_Zs and soc_time_charge_rule_a_Zc, the charge rate is estimated. To make sure the charging time is long enough, we divide the added SOC by a buffer 
hour_start = 0
penalty_delay = 0.05
penalty_fix_cost = 0.01

'''Battery info'''
rate_charge = 0.4
rate_charge_empty = 0.4 # the charge rate when the battery is empty (SOC=0%)
hour_battery_swap = 2 # amount of time required to swap a battery
max_hour_completion = 100 # this is the maximum number of hours allowed from the train departs Sydney until it arrives in Perth.
max_hour_charge = 10  # this is the maximum number of hours allowed to be charged in a row
length_segment_hour = 1 # the length of each time segment when we use piecewise linear approximation to approximate the rate of charge. 0.5 means each segment is 0.5 hours.
number_segments_hour_completion = ceil(max_hour_completion/float(length_segment_hour))  # number of segments over the planning horizon.
number_segments_hour_charge = ceil(max_hour_charge/float(length_segment_hour))  # number of segments over each consecutive charge.
Segments_hour_completion = {}
for l in range(0, number_segments_hour_completion+1):
    Segments_hour_completion.update({l: dict(hour = hour_start+l*length_segment_hour)})
Segments_hour_charge = {}
for l in range(0, number_segments_hour_charge+1):
    Segments_hour_charge.update({l: dict(hour = hour_start+l*length_segment_hour)})
length_segment_SOC = 0.1
number_segments_SOC = ceil(1/float(length_segment_SOC))
Segments_SOC = {}
for l in range(0, number_segments_SOC+1):
    Segments_SOC.update({l: dict(SOC = l*length_segment_SOC)})


for id_instance_random in id_instance_random_list:
    print ('==========================================================')
    print ('Random Instance %d' %id_instance_random)
    id_instance = 'Random Instance %d' %id_instance_random
    current_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = os.path.dirname(current_dir)
    data_file_path = os.path.join(base_dir, 'Data', 'Random', '%s' %id_instance + '.xls')
    result_file_path = os.path.join(base_dir,'Results', 'Random Instances', '%s' %id_instance + '_Delay%d' %(100*penalty_delay) + '_%s' %alg + '.xlsx')
    x2x = XLS2XLSX(data_file_path)
    x2x.to_xlsx(result_file_path)
    
    '''read instance data from excel file'''
    Stations = {}
    Trains = {}
    Containers = {}
    Power = {} # Power[i1][i2] is the amount of power required for a unit-weight train to travel from i1 to i2.
    Power_train = {} # Power_train[j][i1][i2] is the amount of power required for train j to travel from i1 to i2. Power_train[j][i1][i2] = Trains[j]['weight'] * Power[i1][i2]
    TravelTime = {} # TravelTime[i1][i2] is the amount of time required for a new train to travel from i1 to i2.
    TravelTime_train = {}  # TravelTime_train[j][i1][i2] is the amount of time for train j to travel from i1 to i2. It depends on TravelTime[i1][i2] and the age of train j.
    
    databook = xlrd.open_workbook(data_file_path)
    # databook = xlrd.open_workbook('C:\\Jia\My Papers\\FRE\\Data\\Toy Instances\\%s' %id_instance + '.xls')
    datasheet_stations = databook.sheet_by_name("Stations")
    datasheet_trains = databook.sheet_by_name("Trains")
    datasheet_waittime = databook.sheet_by_name("WaitTime")
    for row in range(1, datasheet_stations.nrows):
        i = datasheet_stations.cell_value(row, 0)
        c = datasheet_stations.cell_value(row, 1)
        i_after = datasheet_stations.cell_value(row, 2)
        charger = datasheet_stations.cell_value(row, 3)
        battery = datasheet_stations.cell_value(row, 4)
        Stations.update({i: dict(cost_fix = c, station_after = i_after, \
                                 max_number_batteries = battery, max_number_chargers = charger)})
    for row in range(1, datasheet_trains.nrows):
        j = datasheet_trains.cell_value(row, 0)
        max_bat = int(datasheet_trains.cell_value(row, 1))
        Trains.update({j: dict(containers = [], stations = {})})
        for kid in range(1, max_bat+1):
            k = 'container %d' %kid + ' in %s' %j
            Trains[j]['containers'].append(k)
            Containers.update({k: dict(train = j)})
    for row in range(1, datasheet_waittime.nrows):
        j = datasheet_waittime.cell_value(row, 0)
        for col in range(1, datasheet_waittime.ncols):
            i = datasheet_waittime.cell_value(0, col)
            t = datasheet_waittime.cell_value(row, col)
            Trains[j]['stations'].update({i: dict(time_wait = t)})
    for j in Trains:
        datasheet_power = databook.sheet_by_name("Power_%s" %j)
        datasheet_traveltime = databook.sheet_by_name("TravelTime_%s" %j)
        # read power for train j
        Power_train.update({j: dict()})
        for row in range(1, datasheet_power.nrows):
            i1 = datasheet_power.cell_value(row, 0)
            Power_train[j].update({i1: dict()})
            for col in range(1, datasheet_power.ncols):
                i2 = datasheet_power.cell_value(0, col)
                p = datasheet_power.cell_value(row, col)
                Power_train[j][i1].update({i2: p})
        # read travel time for train j
        TravelTime_train.update({j: dict()})
        for row in range(1, datasheet_traveltime.nrows):
            i1 = datasheet_traveltime.cell_value(row, 0)
            TravelTime_train[j].update({i1: dict()})
            for col in range(1, datasheet_traveltime.ncols):
                i2 = datasheet_traveltime.cell_value(0, col)
                tt = datasheet_traveltime.cell_value(row, col)
                TravelTime_train[j][i1].update({i2: tt})

        

    '''=========================================================================================================
    solve model'''
    time_model_start = time.time()
    # # # solve the constant-rate model
    # # (D, S_arrive, S_depart, T_arrive, T_depart, T_c, X, Y, Z_c, Z_s) \
    # # = Model_change_rate_v2(Stations, Trains, Containers, Power_train, TravelTime_train, hour_battery_swap, rate_charge_empty, penalty_fix_cost, penalty_delay, M)
    
    # solve the changing-rate model
    # (D, S_arrive, S_depart, T_arrive, T_depart, T_c, X, Y, Z_c, Z_s, results) \
    # = Model_change_rate_v2(Stations, Trains, Containers, Power_train, TravelTime_train, hour_battery_swap, rate_charge_empty, penalty_delay, penalty_fix_cost, M, epsilon, buffer_soc_estimate)
    
    # solve by the changng rate model with PLA-rectangle method
    (obj, cost_fix_weight, cost_delay_weight, D, S_arrive, S_depart, T_arrive, T_depart, T_c, X, Y, Z_c, Z_s, F, beta, gamma, tau, eta, B, results, time_model, gap, upper_bound, lower_bound) \
    = Model_change_rate_v2_PLARec(Stations, Trains, Containers, Power_train, TravelTime_train, hour_battery_swap, rate_charge_empty, penalty_delay, penalty_fix_cost, \
                                  M, epsilon, buffer_soc_estimate, Segments_SOC, Segments_hour_charge, gap_op, time_limit_op, mysolver)
    
    time_model_end = time.time()
    time_model = time_model_end - time_model_start
    
    '''=========================================================================================================
    export results to the datafile'''
    # Method 1: copy the xls databook and paste as a xlsx instance book in the same folder so that we can write results in xlsx book 
    # x2x = XLS2XLSX("C:\\Users\\z3547138\\OneDrive - UNSW\\Jia\\Projects\\FRE (Freight Rail Electrification)\\FRE\\Data\\Toy Instances\\%s" %id_instance + ".xls")
    # x2x.to_xlsx("C:\\Users\\z3547138\\OneDrive - UNSW\\Jia\\Projects\\FRE (Freight Rail Electrification)\\FRE\\Data\\Toy Instances\\%s" %id_instance + ".xlsx")
    # x2x = XLS2XLSX("C:\\Jia\My Papers\\FRE\\Data\\Toy Instances\\%s" %id_instance + ".xls")
    # x2x.to_xlsx("C:\\Jia\My Papers\\FRE\\Data\\Toy Instances\\%s" %id_instance + ".xlsx")
    # instancebook = openpyxl.load_workbook('C:\\Users\\z3547138\\OneDrive - UNSW\\Jia\\Projects\\FRE (Freight Rail Electrification)\\FRE\\Data\\Toy Instances\\%s' %id_instance + '.xlsx')
    # instancebook = openpyxl.load_workbook('C:\\Jia\My Papers\\FRE\\Data\\Toy Instances\\%s' %id_instance + '.xlsx')
    
    # Method 2: copy the xls databook and paste as xlsx instance book in another folder under the parent path
    instancebook = openpyxl.load_workbook(result_file_path)
    # work on the instancebook
    instancebook.create_sheet(index = 6, title = "Results_Schedule") 
    schedulesheet = instancebook.get_sheet_by_name("Results_Schedule")
    schedulesheet.cell(row = 1, column = 1).value = 'Trains'
    # build the structure of schedulesheet
    columns = {}  # columns[i] is the column that stores schedules of station i in schedule_sheet
    rows = {} # row[j] or rows[k] is the row that stores schedules of train j or container k in schedule_sheet
    c = 1
    for i in Stations:
        c += 1
        columns.update({i: c})
        schedulesheet.cell(row = 1, column = c).value = i
        if X[i] > 0.5:
            schedulesheet.cell(row = 1, column = c).value = i + ' (deployed)'
    r = 1
    for j in Trains:
        r += 1
        rows.update({j: r})
        rows.update({j + '_T': r+1})
        rows.update({j + '_D': r+2})
        schedulesheet.cell(row = r, column = 1).value = j
        schedulesheet.cell(row = r+1, column = 1).value = '     (arrival time, departure time)'
        schedulesheet.cell(row = r+2, column = 1).value = '     delay (not including scheduled wait time)'
        r += 2
        for k in Trains[j]['containers']:
            r += 1
            rows.update({k: r})
            rows.update({k + '_c/s': r+1})
            rows.update({k + '_S': r+2})
            schedulesheet.cell(row = r, column = 1).value = '     %s' %k
            if Y[j][k] < 0.5:
                schedulesheet.cell(row = r, column = 1).value += ' (no battery)'
            else:
                schedulesheet.cell(row = r, column = 1).value += ' (with battery)'
            schedulesheet.cell(row = r+1, column = 1).value = '            charge? swap?'
            schedulesheet.cell(row = r+2, column = 1).value = '            (arrival SOC, departure SOC)'
            r += 2
    # export statistics to schedulesheet
    for j in Trains:
        for i in Stations:
            schedulesheet.cell(row = rows[j+'_T'], column = columns[i]).value = '(%.2f' %T_arrive[i][j] + ', %.2f' %T_depart[i][j] + ')'
            schedulesheet.cell(row = rows[j+'_D'], column = columns[i]).value = '%.2f' %(D[i][j] - Trains[j]['stations'][i]['time_wait'])
            for k in Trains[j]['containers']:
                if Z_c[i][j][k] > 0.5 and T_c[i][j][k] > epsilon:
                    schedulesheet.cell(row = rows[k+'_c/s'], column = columns[i]).value = 'charge (%.2f' %T_c[i][j][k] + ' hours)'
                elif Z_s[i][j][k] > 0.5:
                    schedulesheet.cell(row = rows[k+'_c/s'], column = columns[i]).value = 'swap'
                schedulesheet.cell(row = rows[k+'_S'], column = columns[i]).value = '(%.2f' %(100*S_arrive[i][j][k]) + '%%, %.2f' %(100*S_depart[i][j][k]) + '%)'
    
    
    # build parameters and conclusion sheets
    instancebook.create_sheet(index = 6, title = "Parameters and Conclusions") 
    conclusionsheet = instancebook.get_sheet_by_name("Parameters and Conclusions")
    conclusionsheet.cell(row = 1, column = 1).value = 'Parameters'
    conclusionsheet.cell(row = 2, column = 1).value = 'Number of stations'
    conclusionsheet.cell(row = 2, column = 2).value = len(Stations)
    conclusionsheet.cell(row = 3, column = 1).value = 'Number of trains'
    conclusionsheet.cell(row = 3, column = 2).value = len(Trains)
    conclusionsheet.cell(row = 4, column = 1).value = 'Average number of batteries each train can carry'
    conclusionsheet.cell(row = 4, column = 2).value = sum(len(Trains[i]['containers']) for i in Trains)/float(len(Trains))
    conclusionsheet.cell(row = 5, column = 1).value = 'Penalty of one unit of fixed cost'
    conclusionsheet.cell(row = 5, column = 2).value = penalty_fix_cost
    conclusionsheet.cell(row = 6, column = 1).value = 'Penalty of one delayed hour'
    conclusionsheet.cell(row = 6, column = 2).value = penalty_delay
    
    conclusionsheet.cell(row = 8, column = 1).value = 'Conclusions'
    conclusionsheet.cell(row = 9, column = 1).value = 'Objective function value'
    obj_ub = penalty_fix_cost * sum((Stations[i]['cost_fix']*X[i]) for i in Stations) \
             + penalty_delay * sum(sum((D[i][j] - Trains[j]['stations'][i]['time_wait']) \
                                       for j in Trains) \
                                   for i in Stations)
    conclusionsheet.cell(row = 9, column = 2).value = obj_ub
    conclusionsheet.cell(row = 10, column = 1).value = 'Number of deployed stations'
    conclusionsheet.cell(row = 10, column = 2).value = sum(X[i] for i in Stations)
    conclusionsheet.cell(row = 11, column = 1).value = 'Fixed cost of deployed stations'
    conclusionsheet.cell(row = 11, column = 2).value = sum((Stations[i]['cost_fix']*X[i]) for i in Stations)
    conclusionsheet.cell(row = 12, column = 1).value = 'Average number of delayed hours for each train (excluding scheduled waiting time)'
    conclusionsheet.cell(row = 12, column = 2).value = sum(sum((D[i][j] - Trains[j]['stations'][i]['time_wait']) \
                                                              for j in Trains) \
                                                          for i in Stations) \
                                                        / len(Trains)
    conclusionsheet.cell(row = 13, column = 1).value = 'Average number of batteries each train carries'
    conclusionsheet.cell(row = 13, column = 2).value = sum(sum(Y[j][k] for k in Trains[j]['containers']) for j in Trains) / float(len(Trains))
    conclusionsheet.cell(row = 14, column = 1).value = 'Average number of hours in charging for each train'
    conclusionsheet.cell(row = 14, column = 2).value = sum(sum(sum(T_c[i][j][k] for k in Trains[j]['containers']) for j in Trains) for i in Stations) / float(len(Trains))
    conclusionsheet.cell(row = 15, column = 1).value = 'Average number of hours in swapping for each train'
    conclusionsheet.cell(row = 15, column = 2).value = hour_battery_swap * sum(sum(sum(Z_s[i][j][k] for k in Trains[j]['containers']) for j in Trains) for i in Stations) / float(len(Trains))
    conclusionsheet.cell(row = 16, column = 1).value = 'Average number of hours in charging for each deployed station'
    conclusionsheet.cell(row = 16, column = 2).value = sum(sum(sum(T_c[i][j][k] for k in Trains[j]['containers']) for j in Trains) for i in Stations) / sum(X[i] for i in Stations)
    conclusionsheet.cell(row = 17, column = 1).value = 'Average number of hours in swapping for each deployed station'
    conclusionsheet.cell(row = 17, column = 2).value = hour_battery_swap * sum(sum(sum(Z_s[i][j][k] for k in Trains[j]['containers']) for j in Trains) for i in Stations) / sum(X[i] for i in Stations)
    conclusionsheet.cell(row = 18, column = 1).value = 'Processing time (sec)'
    conclusionsheet.cell(row = 18, column = 2).value = time_model
    conclusionsheet.cell(row = 19, column = 1).value = 'Lower bound'
    conclusionsheet.cell(row = 19, column = 2).value = lower_bound
    conclusionsheet.cell(row = 20, column = 1).value = 'Optimality gap (%)'
    conclusionsheet.cell(row = 20, column = 2).value = 100 * (obj_ub-lower_bound)/obj_ub
    
    # instancebook.save("C:\\Users\\z3547138\\OneDrive - UNSW\\Jia\\Projects\\FRE (Freight Rail Electrification)\\FRE\\Data\\Toy Instances\\%s" %id_instance + ".xlsx")
    # instancebook.save("C:\\Jia\My Papers\\FRE\\Data\\Toy Instances\\%s" %id_instance + ".xlsx")
    instancebook.save(result_file_path)
    instancebook.close()
    
    
    
    